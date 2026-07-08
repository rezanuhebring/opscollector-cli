from __future__ import annotations

from pathlib import Path

import pytest

from app.services.excel_service import ExcelService


@pytest.fixture()
def svc() -> ExcelService:
    return ExcelService()


class TestExcelService:
    def test_generate_templates(self, svc: ExcelService, tmp_path: Path):
        out = svc.generate_templates(target_dir=tmp_path)
        assert out.exists()
        expected = {
            "template_master.xlsx",
            "template_bau.xlsx",
            "template_okr.xlsx",
            "template_incident.xlsx",
            "template_change.xlsx",
            "template_evidence.xlsx",
        }
        names = {p.name for p in out.iterdir()}
        assert expected.issubset(names)

    def test_preview(self, svc: ExcelService, tmp_path: Path):
        out = svc.generate_templates(target_dir=tmp_path)
        # Add a data row so preview returns something (templates are headers-only).
        from openpyxl import load_workbook

        f = out / "template_bau.xlsx"
        wb = load_workbook(f)
        ws = wb.active
        ws.append(["2026-07-08", "", "Sample Task", "desc", "Completed", "", "", 30, "notes"])
        wb.save(f)
        rows = svc.preview(f)
        assert len(rows) >= 1
        headers = list(rows[0].keys())
        assert headers[0] == "date"

    def test_export_report_bau(self, svc: ExcelService, tmp_path: Path):
        path = svc.export_report("bau", target_dir=tmp_path)
        assert path.exists()
        assert path.name.endswith(".xlsx")

    def test_export_report_okr(self, svc: ExcelService, tmp_path: Path):
        path = svc.export_report("okr", target_dir=tmp_path)
        assert path.exists()

    def test_export_report_incident(self, svc: ExcelService, tmp_path: Path):
        path = svc.export_report("incident", target_dir=tmp_path)
        assert path.exists()

    def test_export_report_evidence(self, svc: ExcelService, tmp_path: Path):
        path = svc.export_report("evidence", target_dir=tmp_path)
        assert path.exists()

    def test_export_report_summary(self, svc: ExcelService, tmp_path: Path):
        path = svc.export_report("summary", target_dir=tmp_path)
        assert path.exists()

    def test_export_report_management(self, svc: ExcelService, tmp_path: Path):
        path = svc.export_report("management", target_dir=tmp_path)
        assert path.exists()

    def test_import_sheet_bau(self, svc: ExcelService, tmp_path: Path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "DailyBAU"
        ws.append(["date", "bau_activity", "title", "description", "status", "pic", "department", "duration_minutes", "notes"])
        ws.append(["2026-07-08", "", "Imported Task", "Imported", "Completed", "", "", 30, "notes"])

        file_path = tmp_path / "bau_import.xlsx"
        wb.save(file_path)

        result = svc.import_sheet(file_path=file_path, entity="daily_bau")
        assert result["created"] == 1
