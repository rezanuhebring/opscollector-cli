"""Excel import/export service.

Import:
  - Reads standard Excel templates for master data, BAU, OKR, Incident, Change.
  - Provides preview (rows without persisting), validation, duplicate detection,
    and transactional import with rollback on failure.

Export:
  - Produces XLSX reports: Daily BAU, Weekly OKR, Incident, Evidence Register,
    Summary Dashboard, Management Report.

This module depends only on the ORM models and openpyxl, so it is reusable by a
future web backend.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.core.config import get_settings
from app.core.exceptions import ImportError_, ValidationError
from app.database.db import get_session
from app.models import (
    BAUActivity,
    ChangeCategory,
    ChangeLog,
    DailyBAU,
    Department,
    Evidence,
    EvidenceCategory,
    Incident,
    IncidentCategory,
    KeyResult,
    Objective,
    OKRProgress,
    PIC,
    Priority,
    Status,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")


# ---------------------------------------------------------------------------
# Validation / reference resolution helpers
# ---------------------------------------------------------------------------

def _require_lookup(session, model, name: str, column: str = "name"):
    """Resolve a reference row by name, raising ValidationError if missing."""
    from sqlalchemy import select
    obj = session.scalars(select(model).where(getattr(model, column) == name)).first()
    if obj is None:
        raise ValidationError(f"{model.__name__} '{name}' not found")
    return obj


def _optional_lookup(session, model, name: str | None, column: str = "name"):
    if not name:
        return None
    from sqlalchemy import select
    return session.scalars(select(model).where(getattr(model, column) == name)).first()


# ---------------------------------------------------------------------------
# IMPORT
# ---------------------------------------------------------------------------

class ExcelService:
    """Excel import and export operations."""

    # --- Template generation ---
    def generate_templates(self, target_dir: str | Path | None = None) -> Path:
        settings = get_settings()
        out = Path(target_dir) if target_dir else settings.templates_dir
        out.mkdir(parents=True, exist_ok=True)
        _write_template(out / "template_master.xlsx", [
            ("Objective", ["name", "title", "owner", "status", "start_date", "end_date", "progress", "description"]),
            ("KeyResult", ["name", "objective", "target_value", "current_value", "unit", "status", "progress", "description"]),
            ("BAUCategory", ["name", "description"]),
            ("BAUActivity", ["name", "bau_category", "description"]),
            ("IncidentCategory", ["name", "description"]),
            ("ChangeCategory", ["name", "description"]),
            ("EvidenceCategory", ["name", "description"]),
            ("Department", ["name", "description"]),
            ("PIC", ["name", "department", "email", "description"]),
            ("Priority", ["name", "level", "description"]),
            ("Status", ["name", "description"]),
        ])
        _write_template(out / "template_bau.xlsx", [
            ("DailyBAU", ["date", "bau_activity", "title", "description", "status", "pic", "department", "duration_minutes", "notes"]),
        ])
        _write_template(out / "template_okr.xlsx", [
            ("OKRProgress", ["key_result", "date", "current_value", "gap", "progress", "achievement", "risk", "issue", "action_plan"]),
        ])
        _write_template(out / "template_incident.xlsx", [
            ("Incident", ["date", "title", "incident_category", "severity", "description", "root_cause", "resolution", "status", "pic", "department", "resolution_time_minutes"]),
        ])
        _write_template(out / "template_change.xlsx", [
            ("ChangeLog", ["date", "title", "change_category", "change_type", "description", "status", "pic", "department", "scheduled_start", "scheduled_end", "result"]),
        ])
        _write_template(out / "template_evidence.xlsx", [
            ("Evidence", ["original_filename", "title", "description", "evidence_category", "entity_type", "entity_id", "uploaded_by"]),
        ])
        return out

    # --- Preview (no persistence) ---
    def preview(self, file_path: str | Path, sheet: str | None = None) -> list[dict[str, Any]]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out = []
        for r in rows[1:]:
            if all(c is None for c in r):
                continue
            out.append({headers[i]: r[i] for i in range(len(headers))})
        return out

    # --- Import with validation + rollback ---
    def import_sheet(
        self,
        file_path: str | Path,
        entity: str,
        *,
        sheet: str | None = None,
        skip_duplicates: bool = True,
    ) -> dict[str, Any]:
        preview = self.preview(file_path, sheet)
        with get_session() as session:
            created = 0
            errors: list[dict[str, Any]] = []
            try:
                for idx, row in enumerate(preview, start=2):  # row 1 = header
                    try:
                        self._import_row(session, entity, row)
                        created += 1
                    except ValidationError as e:
                        if skip_duplicates and "already exists" in str(e):
                            continue
                        errors.append({"row": idx, "error": str(e)})
                        if not skip_duplicates:
                            raise
                if errors and not skip_duplicates:
                    raise ImportError_(f"{len(errors)} row(s) failed validation")
                session.commit()
            except Exception:
                session.rollback()
                raise
        return {"entity": entity, "rows_read": len(preview), "created": created, "errors": errors}

    def _import_row(self, session: Any, entity: str, row: dict[str, Any]) -> None:
        from sqlalchemy import select
        if entity == "daily_bau":
            status = _optional_lookup(session, Status, row.get("status"))
            pic = _optional_lookup(session, PIC, row.get("pic"))
            dept = _optional_lookup(session, Department, row.get("department"))
            activity = _optional_lookup(session, BAUActivity, row.get("bau_activity"))
            session.add(DailyBAU(
                date=str(row.get("date")), title=str(row.get("title")),
                bau_activity_id=activity.id if activity else None,
                description=row.get("description"), status_id=status.id if status else None,
                pic_id=pic.id if pic else None, department_id=dept.id if dept else None,
                duration_minutes=_as_int(row.get("duration_minutes")),
                notes=row.get("notes"),
            ))
        elif entity == "okr_progress":
            kr = _optional_lookup(session, KeyResult, row.get("key_result"))
            if kr is None:
                raise ValidationError("key_result not found")
            session.add(OKRProgress(
                key_result_id=kr.id, date=str(row.get("date")),
                current_value=_as_float(row.get("current_value")),
                gap=_as_float(row.get("gap")), progress=_as_float(row.get("progress")),
                achievement=row.get("achievement"), risk=row.get("risk"),
                issue=row.get("issue"), action_plan=row.get("action_plan"),
            ))
        elif entity == "incident":
            cat = _optional_lookup(session, IncidentCategory, row.get("incident_category"))
            status = _optional_lookup(session, Status, row.get("status"))
            pic = _optional_lookup(session, PIC, row.get("pic"))
            dept = _optional_lookup(session, Department, row.get("department"))
            session.add(Incident(
                date=str(row.get("date")), title=str(row.get("title")),
                incident_category_id=cat.id if cat else None,
                severity=str(row.get("severity") or "Medium"),
                description=row.get("description"), root_cause=row.get("root_cause"),
                resolution=row.get("resolution"), status_id=status.id if status else None,
                pic_id=pic.id if pic else None, department_id=dept.id if dept else None,
                resolution_time_minutes=_as_int(row.get("resolution_time_minutes")),
            ))
        elif entity == "change":
            cat = _optional_lookup(session, ChangeCategory, row.get("change_category"))
            status = _optional_lookup(session, Status, row.get("status"))
            pic = _optional_lookup(session, PIC, row.get("pic"))
            dept = _optional_lookup(session, Department, row.get("department"))
            session.add(ChangeLog(
                date=str(row.get("date")), title=str(row.get("title")),
                change_category_id=cat.id if cat else None,
                change_type=str(row.get("change_type") or "Change"),
                description=row.get("description"), status_id=status.id if status else None,
                pic_id=pic.id if pic else None, department_id=dept.id if dept else None,
                scheduled_start=row.get("scheduled_start"),
                scheduled_end=row.get("scheduled_end"), result=row.get("result"),
            ))
        elif entity == "master":
            # master import is handled per-sheet by intent; default no-op
            raise ValidationError("Use a specific master sheet entity")
        else:
            # Generic master import: entity is a master model key
            from app.services.master_service import MASTER_MODELS
            if entity in MASTER_MODELS:
                model = MASTER_MODELS[entity]
                name = row.get("name")
                if not name:
                    raise ValidationError("name is required")
                existing = session.scalars(select(model).where(model.name == name)).first()
                if existing is not None:
                    raise ValidationError(f"{entity} '{name}' already exists")
                # Only set attributes that are real columns on the model.
                valid_cols = {c.name for c in model.__table__.columns}
                fields = {
                    k: v
                    for k, v in row.items()
                    if k not in ("name",) and v is not None and k in valid_cols
                }
                obj = model(name=name, **fields)  # type: ignore[call-arg]
                session.add(obj)
            else:
                raise ValidationError(f"Unknown import entity '{entity}'")

    # --- EXPORT ---
    def export_report(self, report: str, target_dir: str | Path | None = None) -> Path:
        settings = get_settings()
        out = Path(target_dir) if target_dir else settings.export_dir
        out.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        path = out / f"{report}-{stamp}.xlsx"

        wb = Workbook()
        if report == "bau":
            self._export_bau(wb)
        elif report == "okr":
            self._export_okr(wb)
        elif report == "incident":
            self._export_incident(wb)
        elif report == "evidence":
            self._export_evidence(wb)
        elif report == "summary":
            self._export_summary(wb)
        elif report == "management":
            self._export_management(wb)
        else:
            raise ValidationError(f"Unknown report '{report}'")
        wb.save(path)
        return path

    # export builders
    def _export_bau(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Daily BAU"
        from app.services.bau_service import BAUService
        rows = BAUService().list(limit=None)  # type: ignore[call-arg]
        _sheet_from_rows(ws, ["id", "date", "title", "status", "pic", "department", "duration_minutes", "notes"], rows)

    def _export_okr(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "OKR Progress"
        from app.services.okr_service import OKRService
        rows = OKRService().list(limit=None)  # type: ignore[call-arg]
        _sheet_from_rows(ws, ["id", "key_result", "date", "current_value", "gap", "progress", "achievement", "risk", "issue", "action_plan"], rows)

    def _export_incident(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Incidents"
        from app.services.incident_service import IncidentService
        rows = IncidentService().list(limit=None)  # type: ignore[call-arg]
        _sheet_from_rows(ws, ["id", "incident_no", "date", "title", "severity", "status", "pic", "department", "resolution_time_minutes"], rows)

    def _export_evidence(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Evidence Register"
        from app.services.evidence_service import EvidenceService
        rows = EvidenceService().list(limit=None)  # type: ignore[call-arg]
        _sheet_from_rows(ws, ["id", "original_filename", "stored_filename", "extension", "size_bytes", "entity_type", "entity_id", "title", "uploaded_at"], rows)

    def _export_summary(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Summary Dashboard"
        from app.services.dashboard_service import DashboardService
        s = DashboardService().summary()
        ws.append(["Metric", "Value"])
        for k, v in s.items():
            if isinstance(v, dict):
                for kk, vv in v.items():
                    ws.append([f"{k}.{kk}", vv])
            else:
                ws.append([k, v])
        _style_header(ws, 1)

    def _export_management(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Management Report"
        from app.services.dashboard_service import DashboardService
        ds = DashboardService()
        s = ds.summary()
        ws.append(["OpsCollector Management Report"])
        ws["A1"].font = TITLE_FONT
        ws.append(["Generated", datetime.now().isoformat()])
        ws.append([])
        ws.append(["Section", "Metric", "Value"])
        _style_header(ws, 3)
        for k, v in s.items():
            if isinstance(v, dict):
                for kk, vv in v.items():
                    ws.append(["Objectives" if k == "objectives" else k, kk, vv])
            else:
                ws.append(["KPI", k, v])
        ws.append([])
        ws.append(["Weekly Trend"])
        ws.append(["Week", "BAU", "Incidents", "Changes"])
        _style_header(ws, ws.max_row)
        for w in ds.weekly_trend():
            ws.append([w["week"], w["bau"], w["incidents"], w["changes"]])


# ---------------------------------------------------------------------------
# internal helpers
# ---------------------------------------------------------------------------

def _as_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    return int(v)


def _as_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    return float(v)


def _write_template(path: Path, sheets: list[tuple[str, list[str]]]) -> None:
    wb = Workbook()
    for i, (title, headers) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        ws.append(headers)
        _style_header(ws, 1)
        # example row
        ws.append([""] * len(headers))
    wb.save(path)


def _style_header(ws: Any, row: int) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _sheet_from_rows(ws: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    _style_header(ws, 1)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    # auto width
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(str(h)) + 4))
